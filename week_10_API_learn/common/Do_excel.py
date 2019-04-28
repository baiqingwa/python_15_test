# -*- coding: utf-8 -*-#
# Name:         do_excel
# Author:       白庆
# Email：       364164232@qq.com
# Time:         2019/4/13 0013

from openpyxl import load_workbook
from week_10_API_learn.common import study_session_requests
from week_10_API_learn.common import contents_ml

class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_cases(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_row = sheet.max_row  # 获取最大行数

        cases = []  # 列表，存放所有的测试用例

        for r in range(2, max_row + 1):
            # case = {}
            # case['case_id'] = self.sheet.cell(row=r, column=1)
            # case['title'] = self.sheet.cell(row=r, column=2)

            case = Case()  # 实例
            case.case_id = sheet.cell(row=r, column=1).value
            case.title = sheet.cell(row=r, column=2).value
            case.url = sheet.cell(row=r, column=3).value
            case.data = sheet.cell(row=r, column=4).value
            case.method = sheet.cell(row=r, column=5).value
            case.expected = sheet.cell(row=r, column=6).value
            case.sql = sheet.cell(row=r,column=9).value
            cases.append(case)
        wb.close()
        print(cases) #列表里面存一个一个的实例化对象
        return cases  # 返回case列表

    def write_result(self, row, actual, result):
        wb=load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        wb.save(filename=self.file_name)
        wb.close()

if __name__ == '__main__':
    do_excel=DoExcel(contents_ml.case_file, 'recharge')  #生成类的对象 拿到excel和表单

    cases = do_excel.get_cases() #获取数据 列表存储的实例化对象

    http_request = study_session_requests.HttpRequest_session() #生成完成http请求的对象

    for case in cases:

        # print(case.__dict__)   # ！！！
        resp = http_request.request(case.method, case.url, case.data) #response

        resp_dict = resp.json()['code']  # 返回字典 ！！！
        print(resp_dict)
